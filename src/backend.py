import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import sys

# Read data from an Excel file
def read_excel_file(file_path, sheet_name=0):
    """
    Reads data from an Excel file.

    :param file_path: str, the path to the Excel file
    :param sheet_name: str or int, the sheet name or index (default is 0, the first sheet)
    :return: DataFrame, the data read from the Excel file
    """
    try:
        # Read the Excel file
        data = pd.read_excel(file_path, sheet_name=sheet_name)
        return data
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return None

def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores the path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def vulnDf(df):
    values_to_keep = ['Critical', 'High', 'Medium', 'Low']

    vulnerabilities_df = df[df['Risk'].isin(values_to_keep)]

    return vulnerabilities_df

def essentialDf(df):
    # Column after which to delete all columns
    column_name = 'Plugin Output'

    # Find the index of the specified column
    col_index = df.columns.get_loc(column_name)

    # Keep only the columns up to and including the specified column
    df = df.iloc[:, :col_index + 1]

    return df

def classifiedDf(grouped_df):
    host_groups = {}  # Dictionary to store the grouped DataFrames by host

    for _, group in grouped_df:
        # Get unique pairs of 'Solution' and 'Name'
        unique_pairs = group[['Solution', 'Name']].drop_duplicates()

        for _, pair in unique_pairs.iterrows():
            # Get rows that match the current pair
            matched_group = group[(group['Solution'] == pair['Solution']) & (group['Name'] == pair['Name'])]

            # Group by 'Host' and store in the dictionary
            for host, host_group in matched_group.groupby('Host'):
                if host not in host_groups:
                    host_groups[host] = []
                host_groups[host].append(host_group)

    return host_groups

def removeRedundancy(data_dict):
    updated_dict = {}  # Dictionary to store the updated DataFrames
    
    for host, dfs in data_dict.items():
        # List to store DataFrames with duplicates removed
        cleaned_dfs = []
        
        for df in dfs:
            # Drop duplicate rows from the DataFrame
            cleaned_df = df.drop_duplicates(subset=['Name', 'Solution']).copy()
            
            # Check the 'CVE' column for values
            cve_list = df['CVE'].dropna().tolist()  # Remove NaN values if present

            if len(cve_list) == 0:
                # No CVE numbers present
                cleaned_df['CVE'] = "Not Applicable"
            elif all(isinstance(cve, str) for cve in cve_list):
                # All CVE numbers are strings
                cleaned_df['CVE'] = ', '.join(cve_list)
            
            cleaned_dfs.append(cleaned_df)
        
        # Update the dictionary with the cleaned DataFrames
        updated_dict[host] = cleaned_dfs
    
    return updated_dict


def columnArrangement(new_file_path):
    # Load the existing Excel file into a DataFrame
    df = pd.read_excel(new_file_path)

    # Define a custom sorting order for the 'Risk' column
    risk_order = {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4}

    # Add a new column for sorting purposes based on the custom order
    df['Risk_Order'] = df['Risk'].map(risk_order)

    # Sort the DataFrame by 'Host' and then by the custom 'Risk_Order'
    df = df.sort_values(by=['Host', 'Risk_Order'])

    # Drop the temporary 'Risk_Order' column
    df = df.drop(columns=['Risk_Order'])

    # Add a "Sr No" column at the beginning of the DataFrame
    df.insert(0, 'Sr No', range(1, len(df) + 1))
    # df.insert(6, 'Quick Solution', None) # here ------------------- QS - -----------------------------

    # desired_columns = ['Sr No', 'Risk', 'Host', 'Name', 'Description', 'Solution', 'Quick Solution', 'CVE', 'Port', 'Plugin Output', 'See Also']
    desired_columns = ['Sr No', 'Risk', 'Host', 'Name', 'Description', 'Solution', 'CVE', 'Port', 'Plugin Output', 'See Also']

    # Filter the DataFrame to keep only the desired columns and rearrange them
    df = df[desired_columns]

    # Save the updated DataFrame back to the Excel file, overwriting the existing file
    df.to_excel(new_file_path, index=False)

def coverPage(path):
    workbook = load_workbook(path)

    sheet = workbook.create_sheet(title="CoverPage")
    workbook._sheets.insert(0, workbook._sheets.pop(-1))

    sheet1 = workbook.worksheets[1]
    sheet1.title = "Vulnerabilities Details"

    # workbook.create_sheet(title="Summary").sheet_view.showGridLines = False
    # workbook.create_sheet(title="Overview").sheet_view.showGridLines = False
    # workbook.create_sheet(title="Graph").sheet_view.showGridLines = False

    sheet.sheet_view.showGridLines = False

    sheet.merge_cells('D1:Q10')

    cell = sheet['D1']
    cell.value = "Infrastructure Security\nAssessment Report"

    cell.font = Font(size=54, color="002060", bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # temp_img = resource_path("app_logo.png")
    temp_img = resource_path(os.path.join("docs", "logo", "app_logo.png"))
    img = Image(temp_img)
    img.width = 350
    img.height = 130
    img.anchor = 'D3'

    sheet.add_image(img)

    # Function to add a table with merged header at the top
    def create_table(ws, start_cell, data, table_name, col_widths):
        """
        Adds a table to the worksheet at the specified start cell with given data, merging the first row.

        Parameters:
        ws (Worksheet): The worksheet to add the table to.
        start_cell (str): The top-left cell where the table starts (e.g., "A1").
        data (list): 2D list with the table data (excluding the table name in the first row).
        table_name (str): The name of the table (which will be displayed in the merged first row).
        """
        start_row = ws[start_cell].row
        start_col = ws[start_cell].column

         # Define border style
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # Merge the first row (two columns) and set the table name as the value
        merge_start_cell = ws.cell(row=start_row, column=start_col).coordinate
        merge_end_cell = ws.cell(row=start_row, column=start_col + 1).coordinate
        ws.merge_cells(f"{merge_start_cell}:{merge_end_cell}")
        
        # Set the table name in the merged cell and apply some formatting
        ws[merge_start_cell].value = str(table_name)
        ws[merge_start_cell].alignment = Alignment(horizontal='center')
        ws[merge_start_cell].font = Font(bold=True, size=14)

        ws[merge_start_cell].border = thin_border
        ws[merge_end_cell].border = thin_border

        # Fix the column width for the table columns (separate widths for each column)
        for col, width in zip(range(start_col, start_col + 2), col_widths):
            # col_letter = ws.cell(row=start_row, column=col).column_letter
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width

        # Write the table data starting from the second row (after the merged row)
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                cell = ws.cell(row=start_row + i + 1, column=start_col + j, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    # Sample data for each table (2 columns, 6 rows of content under the header)
    data1 = [["Company", ""], ["Document Title", ""], ["Scope", ""], ["Assesment Date", ""], ["Classification", "Confidential"], ["Document", "Delivarable"]]

    data2 = [["Date", ""], ["Submitted To", ""], ["Designation", ""], ["Address", ""], ["Contact No", ""], ["Email Add", ""]]

    data3 = [["Date", ""], ["Submitted By", ""], ["Designation", ""], ["Address", ""], ["Contact No", ""], ["Email Add", ""]]

    data4 = [["Date", ""], ["Prepared By", ""], ["Designation", ""], ["Address", ""], ["Contact No", ""], ["Email Add", ""]]

    # Add multiple tables with merged headers at different positions
    create_table(sheet, "D12", data1, "Document Authority", col_widths=[15, 30])
    create_table(sheet, "I12", data2, "Document Submission", col_widths=[15, 30])
    create_table(sheet, "N12", data3, "Document Submitted By", col_widths=[15, 30])
    create_table(sheet, "D22", data4, "Document Prepared By", col_widths=[15, 30])

    workbook.save(path)    

def editExcelFile(new_file_path):

    columnArrangement(new_file_path)
    # Define colors for each risk category
    colors = {
        'Critical': 'C00000',  # Red
        'High': 'FF0000',      # Orange
        'Medium': 'FFC000',    # Yellow
        'Low': '00B050',       # Green
    }
    
    # Load the workbook and select the worksheet
    workbook = load_workbook(new_file_path)

    worksheet = workbook['Sheet1']

    worksheet.freeze_panes = 'A2'

    # Find the last column (which contains the "See Also" data)
    last_column = worksheet.max_column  # Assuming "See Also" is the last column

    # Iterate over the rows starting from row 2 (since row 1 is the header)
    for row in worksheet.iter_rows(min_row=2, min_col=last_column, max_col=last_column):
        for cell in row:
            if isinstance(cell.value, str):  # Ensure the cell contains a string
                # Remove hyperlink if any, and force cell content to plain text
                cell.hyperlink = None  # Remove hyperlink
                cell.value = str(cell.value)  # Convert the value to a plain string
    # Define the styles
    header_font = Font(bold=True, size=13, color="FFFFFF")
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")

    # Apply the styles to the header row (row 1)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    worksheet.column_dimensions['A'].width = 7
    worksheet.column_dimensions['B'].width = 9
    worksheet.column_dimensions['C'].width = 14 # Prev = 13
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['E'].width = 40
    worksheet.column_dimensions['F'].width = 40
    # worksheet.column_dimensions['G'].width = 40
    worksheet.column_dimensions['G'].width = 29 # prev = 32
    worksheet.column_dimensions['H'].width = 10
    worksheet.column_dimensions['I'].width = 40
    worksheet.column_dimensions['J'].width = 40

    # # Set the row height to 20 for all rows
    # for row in range(1, worksheet.max_row + 1):
    #     worksheet.row_dimensions[row].height = 30

    # Enable text wrapping with horizontal and vertical alignment for all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    # Iterate through the rows in the specified column (Risk)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
        cell = row[0]  # Get the single cell in the row
        risk_value = cell.value
        if risk_value in colors:
            # Apply the corresponding color to the cell
            cell.fill = PatternFill(start_color=colors[risk_value], end_color=colors[risk_value], fill_type="solid")
        # Save the workbook
        workbook.save(new_file_path)
    # desired_columns = ['D', 'E', 'H', 'I', 'K', 'B', 'G', 'M', 'L']
    coverPage(new_file_path)

def csv_to_excel(csv_file_path):
    # Define the path for the new Excel file
    excel_file_path = os.path.splitext(csv_file_path)[0] + '.xlsx'
    
    # Load the CSV file into a DataFrame
    df = pd.read_csv(csv_file_path)
    
    # Save the DataFrame to an Excel file
    df.to_excel(excel_file_path, index=False)
    
    # Return the path of the newly created Excel file
    return excel_file_path


def main(file_path):
    # Check if the file extension is .csv
    if file_path.lower().endswith('.csv'):
        file_path = csv_to_excel(file_path)    

    sheet_name = 0  # You can specify the sheet name or index here
    df = read_excel_file(file_path, sheet_name)
    
    if df is not None:

        # Removing the redundent columns (All columns are deletd after Plug In Output column)
        essential_df = essentialDf(df)

        # Filtering the points for vulnerability.
        vulnerabilities_df = vulnDf(essential_df)

        group_by_column = ['Solution', 'Name']

        # Group by the specified column
        grouped = vulnerabilities_df.groupby(group_by_column)

        # print("="*100)
        new_groups = classifiedDf(grouped)

        Final_data = removeRedundancy(new_groups)

        pd.set_option('display.max_rows', None)

        # Print the cleaned DataFrames
        # for host, dfs in Final_data.items():
            # print(f"Host: {host}")
            # for i, df in enumerate(dfs):
                # print(f"DataFrame {i + 1}:")
                # print(df)
                # print()
        pd.reset_option('display.max_rows')

        # file_name = os.path.basename(file_path)
        directory, file_name = os.path.split(file_path)
        base_name, ext = os.path.splitext(file_name)
        base_name = base_name + "_converted" + ext
        new_file_path = os.path.join(directory, base_name)

        # Create an ExcelWriter object
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            # Iterate over the dictionary
            all_data = []
            for host, dfs in Final_data.items():
                # Concatenate all DataFrames for the current host
                combined_df = pd.concat(dfs, ignore_index=True)
                all_data.append(combined_df)
            
            final_df = pd.concat(all_data, ignore_index=True)
                
            # Write the combined DataFrame to a single sheet
            final_df.to_excel(writer, sheet_name='All_Hosts', index=False)
        
        editExcelFile(new_file_path)

        return new_file_path
        
    else:
        print("Failed to read the Excel file.")
